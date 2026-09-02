"""Work (作品) routes."""

from flask import Blueprint, request, send_file
from datetime import datetime
from urllib.parse import quote
import zipfile
import os
import tempfile

from ..auth.decorators import login_required, get_current_user
from ..services import work_service, file_service
from ..utils.response import success, error
from ..utils.validators import sanitize_filename_part
from ..config import Config

work_bp = Blueprint('works', __name__, url_prefix='/api/works')


@work_bp.route('', methods=['GET'])
@login_required
def list_works():
    company_id = request.args.get('company_id', type=int)
    agent_id = request.args.get('agent_id', type=int)
    search_keywords = request.args.get('search', '').strip()
    data = work_service.list_works(company_id, agent_id, search_keywords)
    return success(data)


@work_bp.route('/<int:work_id>', methods=['GET'])
@login_required
def get_work(work_id):
    data = work_service.get_work(work_id)
    if not data:
        return error('作品不存在', 404)
    return success(data)


@work_bp.route('/<int:work_id>/history', methods=['GET'])
@login_required
def get_work_history(work_id):
    data = work_service.get_work_history(work_id)
    return success(data)


@work_bp.route('', methods=['POST'])
@login_required
def create_work():
    company_id = request.form.get('company_id', type=int)
    agent_id = request.form.get('agent_id', type=int)
    work_name = request.form.get('work_name', '').strip()
    alias = request.form.get('alias', '').strip()
    proof_file = request.files.get('proof_file')
    other_files = request.files.getlist('other_files')

    if not company_id or not agent_id:
        return error('请选择代理主体和被代理人')
    if not work_name:
        return error('作品名称不能为空')
    if not proof_file or not proof_file.filename:
        return error('请上传权属证明文件')

    safe_work_name = sanitize_filename_part(work_name)

    # 保存权属证明
    proof_filename, proof_warn = file_service.save_proof_file(proof_file, safe_work_name)

    # 保存其他证明
    other_filenames = []
    other_warnings = []
    for idx, f in enumerate([f for f in other_files if f and f.filename], start=1):
        fname, fw = file_service.save_other_proof_file(f, safe_work_name, idx)
        other_filenames.append(fname)
        if fw:
            other_warnings.append(fw)

    data, err = work_service.create_work(
        company_id=company_id,
        agent_id=agent_id,
        work_name=work_name,
        alias=alias if alias else None,
        proof_file=proof_filename,
        other_files=other_filenames if other_filenames else None,
        created_by=get_current_user(),
    )
    if err:
        return error(err)
    warnings = [w for w in [proof_warn] + other_warnings if w]
    return success(data, warning='；'.join(warnings) if warnings else None)


@work_bp.route('/<int:work_id>', methods=['PUT'])
@login_required
def update_work(work_id):
    alias = request.form.get('alias', '').strip() if 'alias' in request.form else None
    proof_file = request.files.get('proof_file')
    other_files = request.files.getlist('other_files')

    existing = work_service.get_work(work_id)
    if not existing:
        return error('作品不存在', 404)

    safe_work_name = sanitize_filename_part(existing['work_name'])

    new_proof = None
    proof_warn = None
    if proof_file and proof_file.filename:
        new_proof, proof_warn = file_service.save_proof_file(proof_file, safe_work_name)

    new_others = None
    other_warnings = []
    valid_others = [f for f in other_files if f and f.filename]
    if valid_others:
        new_others = []
        for idx, f in enumerate(valid_others, start=1):
            fname, fw = file_service.save_other_proof_file(f, safe_work_name, idx)
            new_others.append(fname)
            if fw:
                other_warnings.append(fw)

    data, err = work_service.update_work(
        work_id,
        alias=alias,
        proof_file=new_proof,
        other_files=new_others,
        updated_by=get_current_user()
    )
    if err:
        return error(err)
    warnings = [w for w in [proof_warn] + other_warnings if w]
    return success(data, warning='；'.join(warnings) if warnings else None)


@work_bp.route('/<int:work_id>', methods=['DELETE'])
@login_required
def delete_work(work_id):
    ok, err = work_service.delete_work(work_id)
    if not ok:
        return error(err)
    return success(message='删除成功')


@work_bp.route('/package', methods=['POST'])
@login_required
def package_works():
    data = request.get_json()
    work_ids = data.get('work_ids', [])
    max_size_mb = data.get('max_size_mb', 18)
    excel_mode = data.get('excel_mode', False)
    excel_data = data.get('excel_data', [])
    selected_work_names = data.get('selected_work_names', [])

    if not work_ids:
        return error('请选择要打包的作品')

    # 查询所有作品
    works = []
    for work_id in work_ids:
        work = work_service.get_work(work_id)
        if work:
            works.append(work)

    if not works:
        return error('未找到有效的作品')

    max_size_bytes = max_size_mb * 1024 * 1024  # 转换为字节

    # 收集所有作品的文件及其大小
    work_files = []
    for work in works:
        files = []
        total_size = 0

        # 权属证明
        if work.get('proof_file'):
            proof_path = os.path.join(Config.UPLOAD_FOLDER, '权属证明', work['proof_file'])
            if os.path.exists(proof_path):
                size = os.path.getsize(proof_path)
                files.append({'path': proof_path, 'arcname': work['proof_file'], 'size': size})
                total_size += size

        # 其他证明
        other_files = work.get('other_files', []) or []
        for filename in other_files:
            other_path = os.path.join(Config.UPLOAD_FOLDER, '权属证明', filename)
            if os.path.exists(other_path):
                size = os.path.getsize(other_path)
                files.append({'path': other_path, 'arcname': filename, 'size': size})
                total_size += size

        if files:
            work_files.append({
                'work_name': work['work_name'],
                'files': files,
                'total_size': total_size
            })

    if not work_files:
        return error('没有找到可打包的文件')

    # 分包逻辑：按作品顺序依次放入包，当加入下一个作品会超过限制时，开始新包
    packages = []
    current_package = []
    current_size = 0

    for wf in work_files:
        # 如果当前包为空，或者加入这个作品不超过限制
        if not current_package or (current_size + wf['total_size'] <= max_size_bytes):
            current_package.append(wf)
            current_size += wf['total_size']
        else:
            # 当前包已满，保存并开始新包
            packages.append(current_package)
            current_package = [wf]
            current_size = wf['total_size']

    # 最后一个包
    if current_package:
        packages.append(current_package)

    today = datetime.now().strftime('%Y%m%d')

    # 生成Excel清单函数
    def create_excel_manifest(package_works, package_idx=None):
        """生成Excel清单
        package_works: 当前包中的作品列表
        package_idx: 如果是分包，传入包序号；如果是单包，传None
        """
        import openpyxl
        from openpyxl import Workbook

        if excel_mode and excel_data:
            # Excel模式：提取当前包的作品对应的原始Excel行
            wb = Workbook()
            ws = wb.active
            ws.title = "作品清单"

            # 写入表头（Excel第一行）
            if len(excel_data) > 0:
                header = excel_data[0]
                for col_idx, cell_value in enumerate(header, start=1):
                    ws.cell(row=1, column=col_idx, value=cell_value)

            # 提取当前包的作品名称集合
            package_work_names = set(wf['work_name'] for wf in package_works)

            # 遍历Excel数据行（从第2行开始），筛选匹配的行
            row_num = 2
            for data_row in excel_data[1:]:
                # 检查这一行是否对应当前包中的某个作品
                for cell_value in data_row:
                    if str(cell_value).strip() in package_work_names:
                        # 写入这一行的所有列
                        for col_idx, cell_val in enumerate(data_row, start=1):
                            ws.cell(row=row_num, column=col_idx, value=cell_val)
                        row_num += 1
                        break  # 该行已处理，跳到下一行

            # 保存Excel
            temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb')
            temp_excel.close()
            wb.save(temp_excel.name)
            return temp_excel.name
        else:
            # 搜索栏模式：生成简单的作品名称列表
            wb = Workbook()
            ws = wb.active
            ws.title = "作品清单"

            # 表头
            ws.cell(row=1, column=1, value="作品名称")

            # 数据行
            for idx, wf in enumerate(package_works, start=2):
                ws.cell(row=idx, column=1, value=wf['work_name'])

            # 保存Excel
            temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb')
            temp_excel.close()
            wb.save(temp_excel.name)
            return temp_excel.name

    # 如果只有一个包，直接返回
    if len(packages) == 1:
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip.close()
        temp_excel_path = None

        try:
            # 生成Excel清单
            temp_excel_path = create_excel_manifest(packages[0])

            with zipfile.ZipFile(temp_zip.name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 添加作品文件
                for wf in packages[0]:
                    for file_info in wf['files']:
                        zipf.write(file_info['path'], file_info['arcname'])

                # 添加Excel清单
                zipf.write(temp_excel_path, '作品清单.xlsx')

            zip_filename = f"作品打包_共{len(works)}部_{today}.zip"

            response = send_file(
                temp_zip.name,
                as_attachment=True,
                mimetype='application/zip'
            )

            encoded_filename = quote(zip_filename)
            response.headers['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'

            @response.call_on_close
            def cleanup():
                try:
                    os.unlink(temp_zip.name)
                    if temp_excel_path:
                        os.unlink(temp_excel_path)
                except:
                    pass

            return response
        except Exception as e:
            try:
                os.unlink(temp_zip.name)
                if temp_excel_path:
                    os.unlink(temp_excel_path)
            except:
                pass
            return error(f'打包失败: {str(e)}')

    # 多个包：创建临时目录，生成多个子zip，再打包成总zip
    temp_dir = tempfile.mkdtemp()
    temp_final_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    temp_final_zip.close()

    try:
        sub_zips = []
        for idx, package in enumerate(packages, start=1):
            # 生成当前包的Excel清单
            temp_excel_path = create_excel_manifest(package, idx)

            sub_zip_path = os.path.join(temp_dir, f"作品打包_第{idx}包_共{len(works)}部_{today}.zip")
            with zipfile.ZipFile(sub_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 添加作品文件
                for wf in package:
                    for file_info in wf['files']:
                        zipf.write(file_info['path'], file_info['arcname'])

                # 添加Excel清单
                zipf.write(temp_excel_path, '作品清单.xlsx')

            # 清理临时Excel
            try:
                os.unlink(temp_excel_path)
            except:
                pass

            sub_zips.append(sub_zip_path)

        # 把所有子zip打包到总zip
        with zipfile.ZipFile(temp_final_zip.name, 'w', zipfile.ZIP_DEFLATED) as final_zipf:
            for sub_zip in sub_zips:
                final_zipf.write(sub_zip, os.path.basename(sub_zip))

        final_filename = f"作品打包_共{len(works)}部_分{len(packages)}包_{today}.zip"

        response = send_file(
            temp_final_zip.name,
            as_attachment=True,
            mimetype='application/zip'
        )

        encoded_filename = quote(final_filename)
        response.headers['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'

        @response.call_on_close
        def cleanup():
            try:
                os.unlink(temp_final_zip.name)
                import shutil
                shutil.rmtree(temp_dir, ignore_errors=True)
            except:
                pass

        return response
    except Exception as e:
        try:
            os.unlink(temp_final_zip.name)
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass
        return error(f'打包失败: {str(e)}')

